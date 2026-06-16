# backend/app/api/v1/admin_faculty.py

from fastapi import (
    APIRouter,
    Depends,
    HTTPException,
    status,
)
from sqlalchemy.orm import Session
from fastapi import UploadFile, File
from openpyxl import load_workbook
from io import BytesIO

from app.schemas.admin_import import (
    BulkImportResponse,
)

from app.core.constants.roles import UserRole
from app.core.dependencies import require_admin
from app.core.security import get_password_hash
from app.db.session import get_db
from app.models.user import User
from app.schemas.admin_faculty import (
    FacultyCreate,
    FacultyUpdate,
    FacultyResponse,
)

router = APIRouter(
    prefix="/admin/faculty",
    tags=["Admin Faculty"],
    dependencies=[Depends(require_admin)],
)


@router.post(
    "",
    response_model=FacultyResponse,
    status_code=status.HTTP_201_CREATED,
)
def create_faculty(
    data: FacultyCreate,
    db: Session = Depends(get_db),
):
    email = data.email.lower().strip()

    existing_faculty = (
        db.query(User)
        .filter(
            User.email == email,
        )
        .first()
    )

    if existing_faculty:
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail="Email already exists",
        )

    faculty = User(
        full_name=data.full_name.strip(),
        email=email,
        password_hash=get_password_hash(
            data.password,
        ),
        role=UserRole.FACULTY.value,
        is_active=True,
    )

    db.add(faculty)
    db.commit()
    db.refresh(faculty)

    return faculty


@router.post(
    "/import",
    response_model=BulkImportResponse,
)
def import_faculty(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
):
    # Fix 3: Validate file type before processing
    if not file.filename.endswith((".xlsx", ".xlsm")):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Excel file required (.xlsx or .xlsm)",
        )

    created = 0
    skipped = 0
    errors: list[str] = []

    try:
        workbook = load_workbook(
            filename=BytesIO(
                file.file.read(),
            ),
        )

        worksheet = workbook.active

    except Exception:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Invalid Excel file",
        )

    for row_number, row in enumerate(
        worksheet.iter_rows(
            min_row=2,
            values_only=True,
        ),
        start=2,
    ):
        try:
            # Fix 6: Skip truly empty rows (handles (None, None, None) case)
            if all(cell is None for cell in row):
                skipped += 1
                continue

            # Fix 1: Validate minimum column count
            if len(row) < 4:
                skipped += 1
                errors.append(
                    f"Row {row_number}: Expected 4 columns (ID, Name, Email, Password)",
                )
                continue

            faculty_id = row[0]

            full_name = (
                str(row[1]).strip()
                if row[1] is not None
                else ""
            )

            email = (
                str(row[2]).strip().lower()
                if row[2] is not None
                else ""
            )

            password = (
                str(row[3]).strip()
                if row[3] is not None
                else ""
            )

            if not full_name:
                skipped += 1
                errors.append(
                    f"Row {row_number}: Missing name",
                )
                continue

            if not email:
                skipped += 1
                errors.append(
                    f"Row {row_number}: Missing email",
                )
                continue

            if not password:
                skipped += 1
                errors.append(
                    f"Row {row_number}: Missing password",
                )
                continue

            existing_faculty = (
                db.query(User)
                .filter(
                    User.email == email,
                )
                .first()
            )

            # Fix 4: Report duplicate email reason
            if existing_faculty:
                skipped += 1
                errors.append(
                    f"Row {row_number}: Email already exists ({email})",
                )
                continue

            faculty = User(
                full_name=full_name,
                email=email,
                password_hash=get_password_hash(
                    password,
                ),
                role=UserRole.FACULTY.value,
                is_active=True,
            )

            # Fix 2: Flush per row to catch DB constraint errors early
            try:
                db.add(faculty)
                db.flush()
                created += 1
            except Exception as exc:
                db.rollback()
                skipped += 1
                errors.append(
                    f"Row {row_number}: Database error — {str(exc)}",
                )

        except Exception as exc:
            skipped += 1
            errors.append(
                f"Row {row_number}: {str(exc)}",
            )

    db.commit()

    return {
        "created": created,
        "skipped": skipped,
        "errors": errors,
    }


@router.get(
    "",
    response_model=list[FacultyResponse],
)
def get_faculty(
    db: Session = Depends(get_db),
):
    faculty = (
        db.query(User)
        .filter(
            User.role == UserRole.FACULTY.value,
        )
        .order_by(
            User.id.asc(),
        )
        .all()
    )

    return faculty


@router.get(
    "/{faculty_id}",
    response_model=FacultyResponse,
)
def get_faculty_member(
    faculty_id: int,
    db: Session = Depends(get_db),
):
    faculty = (
        db.query(User)
        .filter(
            User.id == faculty_id,
            User.role == UserRole.FACULTY.value,
        )
        .first()
    )

    if not faculty:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Faculty not found",
        )

    return faculty


@router.put(
    "/{faculty_id}",
    response_model=FacultyResponse,
)
def update_faculty(
    faculty_id: int,
    data: FacultyUpdate,
    db: Session = Depends(get_db),
):
    faculty = (
        db.query(User)
        .filter(
            User.id == faculty_id,
            User.role == UserRole.FACULTY.value,
        )
        .first()
    )

    if not faculty:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Faculty not found",
        )

    email = data.email.lower().strip()

    existing_faculty = (
        db.query(User)
        .filter(
            User.email == email,
            User.id != faculty_id,
        )
        .first()
    )

    if existing_faculty:
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail="Email already exists",
        )

    faculty.full_name = data.full_name.strip()
    faculty.email = email
    faculty.is_active = data.is_active

    db.commit()
    db.refresh(faculty)

    return faculty


@router.delete(
    "/{faculty_id}",
    status_code=status.HTTP_200_OK,
)
def delete_faculty(
    faculty_id: int,
    db: Session = Depends(get_db),
):
    faculty = (
        db.query(User)
        .filter(
            User.id == faculty_id,
            User.role == UserRole.FACULTY.value,
        )
        .first()
    )

    if not faculty:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Faculty not found",
        )

    db.delete(faculty)
    db.commit()

    return {
        "message": "Faculty deleted successfully",
    }