# backend/app/api/v1/admin_students.py

from fastapi import (
    APIRouter,
    Depends,
    HTTPException,
    status,
)
from sqlalchemy.orm import Session

from fastapi import UploadFile, File
from openpyxl import load_workbook
from io import BytesIO

from app.schemas.admin_import import (
    BulkImportResponse,
)

from app.core.constants.roles import UserRole
from app.core.dependencies import require_admin
from app.core.security import get_password_hash
from app.db.session import get_db
from app.models.user import User
from app.models.device import Device
from app.models.auth_session import AuthSession
from app.schemas.admin_student import (
    StudentCreate,
    StudentUpdate,
    StudentResponse,
)

router = APIRouter(
    prefix="/admin/students",
    tags=["Admin Students"],
    dependencies=[Depends(require_admin)],
)


@router.post(
    "",
    response_model=StudentResponse,
    status_code=status.HTTP_201_CREATED,
)
def create_student(
    data: StudentCreate,
    db: Session = Depends(get_db),
):
    email = data.email.lower().strip()

    existing_student = (
        db.query(User)
        .filter(
            User.email == email,
        )
        .first()
    )

    if existing_student:
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail="Email already exists",
        )

    student = User(
        full_name=data.full_name.strip(),
        email=email,
        password_hash=get_password_hash(
            data.password,
        ),
        role=UserRole.STUDENT.value,
        is_active=True,
    )

    db.add(student)
    db.commit()
    db.refresh(student)

    return student


@router.post(
    "/import",
    response_model=BulkImportResponse,
)
def import_students(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
):
    # Fix 3: Validate file type before processing
    if not file.filename.endswith((".xlsx", ".xlsm")):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Excel file required (.xlsx or .xlsm)",
        )

    created = 0
    skipped = 0
    errors: list[str] = []

    try:
        workbook = load_workbook(
            filename=BytesIO(
                file.file.read(),
            ),
        )

        worksheet = workbook.active

    except Exception:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Invalid Excel file",
        )

    for row_number, row in enumerate(
        worksheet.iter_rows(
            min_row=2,
            values_only=True,
        ),
        start=2,
    ):
        try:
            # Fix 6: Skip truly empty rows (handles (None, None, None) case)
            if all(cell is None for cell in row):
                skipped += 1
                continue

            # Fix 1: Validate minimum column count
            if len(row) < 4:
                skipped += 1
                errors.append(
                    f"Row {row_number}: Expected 4 columns (ID, Name, Email, Password)",
                )
                continue

            student_id = row[0]

            full_name = (
                str(row[1]).strip()
                if row[1] is not None
                else ""
            )

            email = (
                str(row[2]).strip().lower()
                if row[2] is not None
                else ""
            )

            password = (
                str(row[3]).strip()
                if row[3] is not None
                else ""
            )

            if not full_name:
                skipped += 1
                errors.append(
                    f"Row {row_number}: Missing name",
                )
                continue

            if not email:
                skipped += 1
                errors.append(
                    f"Row {row_number}: Missing email",
                )
                continue

            if not password:
                skipped += 1
                errors.append(
                    f"Row {row_number}: Missing password",
                )
                continue

            existing_student = (
                db.query(User)
                .filter(
                    User.email == email,
                )
                .first()
            )

            # Fix 4: Report duplicate email reason
            if existing_student:
                skipped += 1
                errors.append(
                    f"Row {row_number}: Email already exists ({email})",
                )
                continue

            student = User(
                full_name=full_name,
                email=email,
                password_hash=get_password_hash(
                    password,
                ),
                role=UserRole.STUDENT.value,
                is_active=True,
            )

            # Fix 2: Flush per row to catch DB constraint errors early
            try:
                db.add(student)
                db.flush()
                created += 1
            except Exception as exc:
                db.rollback()
                skipped += 1
                errors.append(
                    f"Row {row_number}: Database error — {str(exc)}",
                )

        except Exception as exc:
            skipped += 1
            errors.append(
                f"Row {row_number}: {str(exc)}",
            )

    db.commit()

    return {
        "created": created,
        "skipped": skipped,
        "errors": errors,
    }


@router.get(
    "",
    response_model=list[StudentResponse],
)
def get_students(
    is_active: bool | None = None,
    db: Session = Depends(get_db),
):
    query = (
        db.query(User)
        .filter(
            User.role == UserRole.STUDENT.value,
        )
    )

    if is_active is not None:
        query = query.filter(
            User.is_active == is_active,
        )

    students = (
        query
        .order_by(User.id.asc())
        .all()
    )

    return students


@router.get(
    "/{student_id}",
    response_model=StudentResponse,
)
def get_student(
    student_id: int,
    db: Session = Depends(get_db),
):
    student = (
        db.query(User)
        .filter(
            User.id == student_id,
            # Fix 5: Use enum consistently
            User.role == UserRole.STUDENT.value,
        )
        .first()
    )

    if not student:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Student not found",
        )

    return student


@router.put(
    "/{student_id}",
    response_model=StudentResponse,
)
def update_student(
    student_id: int,
    data: StudentUpdate,
    db: Session = Depends(get_db),
):
    student = (
        db.query(User)
        .filter(
            User.id == student_id,
            # Fix 5: Use enum consistently
            User.role == UserRole.STUDENT.value,
        )
        .first()
    )

    if not student:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Student not found",
        )

    email = data.email.lower().strip()

    existing_student = (
        db.query(User)
        .filter(
            User.email == email,
            User.id != student_id,
        )
        .first()
    )

    if existing_student:
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail="Email already exists",
        )

    student.full_name = data.full_name.strip()
    student.email = email
    student.is_active = data.is_active

    db.commit()
    db.refresh(student)

    return student


@router.delete(
    "/{student_id}",
    status_code=status.HTTP_200_OK,
)
def delete_student(
    student_id: int,
    db: Session = Depends(get_db),
):
    student = (
        db.query(User)
        .filter(
            User.id == student_id,
            User.role == UserRole.STUDENT.value,
        )
        .first()
    )

    if not student:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Student not found",
        )

    device_ids = [
        device.id
        for device in db.query(Device)
        .filter(
            Device.user_id == student.id,
        )
        .all()
    ]

    db.query(AuthSession).filter(
        AuthSession.user_id == student.id,
    ).delete(
        synchronize_session=False,
    )

    if device_ids:
        db.query(AuthSession).filter(
            AuthSession.device_id.in_(device_ids),
        ).delete(
            synchronize_session=False,
        )

    db.query(Device).filter(
        Device.user_id == student.id,
    ).delete(
        synchronize_session=False,
    )

    student.is_active = False

    db.commit()
    db.refresh(student)

    return {
        "message": "Student deactivated successfully",
    }


@router.post(
    "/{student_id}/activate",
    status_code=status.HTTP_200_OK,
)
def activate_student(
    student_id: int,
    db: Session = Depends(get_db),
):
    student = (
        db.query(User)
        .filter(
            User.id == student_id,
            User.role == UserRole.STUDENT.value,
        )
        .first()
    )

    if not student:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Student not found",
        )

    student.is_active = True

    db.commit()
    db.refresh(student)

    return {
        "message": "Student activated successfully",
    }