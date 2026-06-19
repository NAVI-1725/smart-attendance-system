# backend/app/services/beacon_import_service.py

from io import BytesIO

from fastapi import HTTPException, UploadFile, status
from openpyxl import load_workbook
from sqlalchemy.orm import Session
from sqlalchemy.exc import IntegrityError

from app.models.classroom import Classroom
from app.models.trusted_ble_beacon import TrustedBLEBeacon
from app.schemas.beacon_import import (
    BeaconImportError,
    BeaconImportResult,
)


def import_beacons_from_excel(
    db: Session,
    file: UploadFile,
) -> BeaconImportResult:
    if not file.filename:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="No file provided",
        )

    if not file.filename.endswith(".xlsx"):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Only .xlsx files are supported",
        )

    try:
        workbook = load_workbook(
            filename=BytesIO(file.file.read()),
            data_only=True,
        )
    except Exception:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Invalid Excel file",
        )

    worksheet = workbook.active

    created = 0
    skipped = 0
    errors: list[BeaconImportError] = []

    seen_uuids: set[str] = set()

    for row_number, row in enumerate(
        worksheet.iter_rows(
            min_row=2,
            values_only=True,
        ),
        start=2,
    ):
        classroom_id = row[0]
        beacon_uuid = row[1]
        beacon_name = row[2]

        normalized_uuid = (
            str(beacon_uuid).strip()
            if beacon_uuid is not None
            else ""
        )

        if (
            classroom_id is None
            or beacon_uuid is None
        ):
            skipped += 1

            errors.append(
                BeaconImportError(
                    row=row_number,
                    reason=(
                        "classroom_id and "
                        "beacon_uuid are required"
                    ),
                )
            )

            continue

        if not normalized_uuid:
            skipped += 1

            errors.append(
                BeaconImportError(
                    row=row_number,
                    reason="Beacon UUID cannot be empty",
                )
            )

            continue

        if normalized_uuid in seen_uuids:
            skipped += 1

            errors.append(
                BeaconImportError(
                    row=row_number,
                    reason="Duplicate UUID in Excel file",
                )
            )

            continue

        classroom = (
            db.query(Classroom)
            .filter(
                Classroom.id == classroom_id,
            )
            .first()
        )

        if not classroom:
            skipped += 1

            errors.append(
                BeaconImportError(
                    row=row_number,
                    reason=(
                        f"Classroom "
                        f"{classroom_id} "
                        f"does not exist"
                    ),
                )
            )

            continue

        existing_beacon = (
            db.query(TrustedBLEBeacon)
            .filter(
                TrustedBLEBeacon.beacon_uuid
                == normalized_uuid
            )
            .first()
        )

        if existing_beacon:
            skipped += 1

            errors.append(
                BeaconImportError(
                    row=row_number,
                    reason=(
                        f"Beacon UUID "
                        f"'{beacon_uuid}' "
                        f"already exists"
                    ),
                )
            )

            continue

        beacon = TrustedBLEBeacon(
            classroom_id=int(classroom_id),
            beacon_uuid=normalized_uuid,
            beacon_name=(
                str(beacon_name).strip()
                if beacon_name is not None
                else None
            ),
        )

        seen_uuids.add(
            normalized_uuid,
        )

        db.add(beacon)

        created += 1

    try:
        db.commit()
    except IntegrityError:
        db.rollback()

        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=(
                "Beacon import failed due to "
                "duplicate or invalid data"
            ),
        )

    return BeaconImportResult(
        created=created,
        skipped=skipped,
        errors=errors,
    )